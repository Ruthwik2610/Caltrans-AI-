import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO

def create_excel(file_path):
    # --- Data for Sheet 1: Feature Implementation Status ---
    features_data = [
        ["Core Evaluation Engine", "Implemented", "GPT-4o based evaluation of 25-question rubric.", "Currently uses single document input (Fact Sheet)."],
        ["Rubric Question Mapping", "Implemented", "25 questions from Tables 3-8 mapped with ABC options.", "Weights are applied based on Caltrans standards."],
        ["Scoring & Recommendation", "Implemented", "Weighted composite score calculation and method recommendation.", "Supports borderline case detection."],
        ["Override Logic", "Implemented", "9 document-based override rules (e.g., A1=C blocks DBB).", "Transparently logged in recommendation logic."],
        ["Excel Export", "Implemented", "Styled Excel output with results, recommendations, and completeness.", "Need to enhance formatting for 'Professional yet Elegant' look."],
        ["Multi-Method Scoring (3.1)", "Planned", "Display scores for all methods (DBB, DB, CMGC, PDB).", "Requires update to scoring matrix to generate parallel scores."],
        ["Detailed Borderline Analysis (3.2)", "In Progress", "Pros/Cons for top 2-3 methods when scores are close.", "Need to refine the 'influencing factors' extraction logic."],
        ["Validation Mode (3.3)", "In Progress", "Compare AI ratings vs District pre-filled ratings.", "Basic extraction exists; flagging mismatches needs UI/Excel integration."],
        ["Multi-Document Support (3.6)", "Planned", "Cross-validation using project reports and additional PDFs.", "Requires MCP server or RAG-based multi-doc context handling."],
        ["AI vs Expected Analysis (3.7)", "Planned", "Identify questions causing deviation from human result.", "Will require a 'ground truth' input for comparison."],
        ["Output Customization (3.10)", "Planned", "Tailored views for District teams vs Internal reviewers.", "Targeting 'Guidance' for Districts and 'Full Comparison' for HQ."],
    ]
    df_features = pd.DataFrame(features_data, columns=["Feature/Function", "Status", "Description", "Technical Remarks / Implementation Details"])

    # --- Data for Sheet 2: Feedback Alignment Matrix ---
    feedback_data = [
        ["3.1 Multi-Method Output", "High", "Implemented runner-up detection; full list planned.", "Will add a 'Comparison Matrix' sheet to the output Excel."],
        ["3.2 Borderline Case Handling", "High", "Basic borderline flag exists.", "Developing 'Influencing Factors' summary for top 3 candidates."],
        ["3.3 Validation Mode", "Very Important", "Can extract pre-filled ratings.", "Need to implement 'Mismatched Justification' flagging."],
        ["3.4 Missing Info Detection", "High", "Implemented basic detection.", "Enhancing to show 'Confidence Impact' per missing field."],
        ["3.6 Multi-Document Support", "Important", "Single-doc focused currently.", "Planning integration with local 'Institutional Memory' and Project Reports."],
        ["3.9 Target User Alignment", "Critical", "Focusing on HQ review.", "Shifting toward 'Guidance' for Districts to reduce subjectivity."],
    ]
    df_feedback = pd.DataFrame(feedback_data, columns=["Requirement", "Priority", "Current Alignment", "Next Steps"])

    # --- Data for Sheet 3: Doubts & Clarifications ---
    doubts_data = [
        ["Scoring Weights", "Are the current section weights (A=25%, B=20%, etc.) finalized or should they be adjustable by HQ?", "Technical"],
        ["Ground Truth", "Do we have a set of 'Expected Results' (Human evaluations) to test the AI vs Expected Analysis?", "Data"],
        ["Multi-Doc Scope", "Which 'Additional Supporting Documents' are most critical for validation? (e.g., PSR, PDS, Environmental Reports?)", "Process"],
        ["Override Rules", "Are there any new 'Legislative' or 'Policy' overrides not covered in the 9 current rules?", "Policy"],
        ["User Interface", "Should 'Validation Mode' be an interactive step (User corrects AI) or a static report comparison?", "UX"],
    ]
    df_doubts = pd.DataFrame(doubts_data, columns=["Area of Concern", "Doubt / Question for Clarification", "Category"])

    # Create Excel writer
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        df_features.to_excel(writer, sheet_name='Implementation Progress', index=False)
        df_feedback.to_excel(writer, sheet_name='Feedback Alignment', index=False)
        df_doubts.to_excel(writer, sheet_name='Doubts & Clarifications', index=False)

    # --- Styling with Openpyxl ---
    wb = openpyxl.load_workbook(file_path)
    
    # Styles
    header_font = Font(name='Segoe UI', bold=True, color='FFFFFF', size=12)
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    cell_font = Font(name='Segoe UI', size=10)
    alignment_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    alignment_left = Alignment(horizontal='left', vertical='top', wrap_text=True)
    border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    status_colors = {
        "Implemented": PatternFill(start_color='C6E0B4', end_color='C6E0B4', fill_type='solid'), # Green
        "In Progress": PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid'), # Yellow
        "Planned": PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid'),    # Grey
    }

    priority_colors = {
        "High": Font(color='C00000', bold=True),
        "Very Important": Font(color='C00000', bold=True),
        "Critical": Font(color='C00000', bold=True, italic=True),
    }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Style Headers
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment_center
            cell.border = border_thin
        
        # Style Data Cells
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = cell_font
                cell.alignment = alignment_left
                cell.border = border_thin
                
                # Conditional formatting for Status
                if sheet_name == 'Implementation Progress' and cell.column == 2:
                    if cell.value in status_colors:
                        cell.fill = status_colors[cell.value]
                        cell.alignment = alignment_center
                
                # Conditional formatting for Priority
                if sheet_name == 'Feedback Alignment' and cell.column == 2:
                    if cell.value in priority_colors:
                        cell.font = priority_colors[cell.value]
                        cell.alignment = alignment_center

        # Adjust Column Widths
        column_widths = {
            'Implementation Progress': [30, 15, 50, 60],
            'Feedback Alignment': [30, 15, 50, 50],
            'Doubts & Clarifications': [25, 60, 15]
        }
        
        for i, width in enumerate(column_widths.get(sheet_name, []), start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    wb.save(file_path)

if __name__ == "__main__":
    output_path = "Project_Delivery_Evaluator_Progress_Report.xlsx"
    create_excel(output_path)
    print(f"Excel created at: {output_path}")
